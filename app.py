import io
import requests
import pandas as pd
import openpyxl
import streamlit as st
import premailer
from openpyxl.utils import get_column_letter

# ─── Helpers ────────────────────────────────────────────────────────────────────
def format_value(val, fmt):
    if val is None:
        return ""
    try:
        if "£" in fmt or "\u00a3" in fmt:
            return f"£{float(val):,.2f}"
        elif "€" in fmt or "\u20ac" in fmt:
            return f"€{float(val):,.2f}"
        elif "$" in fmt or "\u0024" in fmt:
            return f"${float(val):,.2f}"
        if isinstance(val, float):
            return f"{val:,.2f}"
    except:
        pass
    return str(val)

def url_ok(url):
    try:
        r = requests.head(url, allow_redirects=True, timeout=2)
        return r.status_code < 400
    except:
        return False

# ─── 1. Page config & load mapping ──────────────────────────────────────────────
st.set_page_config(layout="wide", page_title="Offer → Email-HTML")
st.title("Intamarques Offer to Email HTML Converter")

# Load mapping.csv
mapping_path = "mapping.csv"
try:
    mapping_df = pd.read_csv(mapping_path)
except Exception as e:
    st.error(f"⚠️ Could not read {mapping_path}: {e}")
    st.stop()
if not {"Code","Image URL"}.issubset(mapping_df.columns):
    st.error("⚠️ mapping.csv must have 'Code' and 'Image URL' columns")
    st.stop()
mapping_df = mapping_df.drop_duplicates(subset=["Code"], keep="first")

# ─── 2. Upload offer sheet ─────────────────────────────────────────────────────
uploaded = st.file_uploader("1) Upload your Offer .xlsx", type="xlsx")
if not uploaded:
    st.info("Please upload your offer file to proceed.")
    st.stop()
wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
sheet_name = st.selectbox("2) Select sheet", wb.sheetnames)
ws = wb[sheet_name]

# ─── 3. Auto-detect header row ───────────────────────────────────────────────────
auto_header = 1
for i in range(1, 21):
    vals = [c.value for c in ws[i]]
    if sum(1 for v in vals if v not in (None, "")) > len(vals) / 2:
        auto_header = i
        break
header_row = st.number_input(
    "3) Header row (detected); use +/− to adjust (usually 6):",
    min_value=1, max_value=20, value=auto_header
)

# ─── 4. Read and clean data ─────────────────────────────────────────────────────
buf = io.BytesIO(uploaded.getvalue())
df = pd.read_excel(buf, sheet_name=sheet_name, header=header_row-1, engine="openpyxl")
df.dropna(how="all", axis=0, inplace=True)
df.dropna(how="all", axis=1, inplace=True)
df = df.loc[:, ~df.columns.str.match(r"^Unnamed")]
st.subheader("4) Raw Offer Data")
st.dataframe(df, use_container_width=True)

# ─── 5. Pick columns ────────────────────────────────────────────────────────────
cols = st.multiselect(
    "5) Columns to include in email — uncheck those you don't need:",
    options=list(df.columns),
    default=list(df.columns)
)
if not cols:
    st.warning("Select at least one column.")
    st.stop()
df_view = df[cols]

# ─── 6. Include product images? ─────────────────────────────────────────────────
use_images = st.checkbox("Include product images?")
if use_images:
    merged = pd.merge(
        df_view,
        mapping_df[["Code", "Image URL"]],
        how="left",
        on="Code",
        validate="many_to_one"
    )
else:
    merged = df_view.copy()
    merged["Image URL"] = ""

# ─── 7. Build preview rows ──────────────────────────────────────────────────────
excel_headers = [c.value for c in ws[header_row]]
preview_rows = []
for idx, row in enumerate(merged.itertuples(index=False, name=None)):
    excel_r = header_row + 1 + idx
    img_html = ""
    if use_images:
        url = row[len(cols)]
        if isinstance(url, str) and url_ok(url):
            img_html = f'<img src="{url}" style="height:60px; width:auto;" />'
    cells = [img_html] if use_images else []
    for j, col in enumerate(cols):
        cell = ws.cell(row=excel_r, column=excel_headers.index(col)+1)
        cells.append(format_value(cell.value, cell.number_format or ""))
    preview_rows.append(cells)
preview_cols = (['Image'] if use_images else []) + cols
preview_df = pd.DataFrame(preview_rows, columns=preview_cols)
sub_header = "7) Preview with Images" if use_images else "7) Preview"
st.subheader(sub_header)
st.write(preview_df.to_html(escape=False, index=False), unsafe_allow_html=True)

# ─── 8. Header & column colors ─────────────────────────────────────────────────
st.markdown("---")
st.subheader("8) Choose colors for headers/columns")
hdr_cols = st.multiselect("Select headers/columns to color:", preview_cols)
col_mapping = {c: st.color_picker(f"Color for {c}", "#f0f0f0", key=c) for c in hdr_cols}

# ─── 9. Generate Brevo HTML ────────────────────────────────────────────────────
if st.button("9) Generate Brevo-Ready HTML"):
    def build_html():
        # Generate base HTML with pandas
        raw = preview_df.to_html(
            escape=False,
            index=False,
            table_id="offer-table",
            classes=["offer-table"]
        )
        # Build CSS
        css = """
<style>
#offer-table { border-collapse:collapse; font-family:Arial,sans-serif; font-size:12px; table-layout:auto; width:auto; }
#offer-table th, #offer-table td { border:1px solid #ccc; padding:6px; text-align:center; white-space:normal; word-wrap:break-word; }
#offer-table th { background-color:#f0f0f0; font-weight:bold; word-break:keep-all; }
"""
        # Column-specific colors
        for col, color in col_mapping.items():
            idx = preview_cols.index(col) + 1
            css += f"\n#offer-table th:nth-child({idx}), #offer-table td:nth-child({idx}) {{ background-color: {color}; }}"
        css += "\n</style>"
        return premailer.transform(css + raw)

    email_html = build_html()
    st.subheader("📧 Your Brevo-Ready HTML")
    st.components.v1.html(email_html, height=600, scrolling=True)
    st.text_area("Copy this HTML:", email_html, height=300)


